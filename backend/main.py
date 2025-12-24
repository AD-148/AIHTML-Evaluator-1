from fastapi import FastAPI, HTTPException, UploadFile, File
from pydantic import BaseModel
from typing import Optional, List
from fastapi.middleware.cors import CORSMiddleware
import os
import asyncio
from starlette.concurrency import run_in_threadpool
from dotenv import load_dotenv

try:
    from backend.llm_service import analyze_chat, EvaluationResult
except ImportError:
    try:
        from .llm_service import analyze_chat, EvaluationResult
    except ImportError:
         from llm_service import analyze_chat, EvaluationResult

load_dotenv()

app = FastAPI(title="HTML LLM Judge")

# Configure CORS for frontend
# In production, Nginx handles requests as same-origin.
# These origins are mainly for local development.
origins = [
    "http://localhost:5173",
    "http://127.0.0.1:5173",
    "http://localhost:3000",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class Message(BaseModel):
    role: str
    content: str

class ChatInput(BaseModel):
    messages: List[Message]
    model_provider: Optional[str] = "openai"

@app.get("/")
def read_root():
    return {"message": "HTML LLM Judge API is running"}

@app.post("/evaluate", response_model=EvaluationResult)
async def evaluate_conversation(input_data: ChatInput):
    if not input_data.messages:
        raise HTTPException(status_code=400, detail="Messages cannot be empty")
    
    try:
        # Convert Pydantic models to dicts
        msgs = [m.model_dump() for m in input_data.messages]
        result = await analyze_chat(msgs)
        return result
    except Exception as e:
        logger.error(f"Error evaluating conversation: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/batch/process")
async def batch_evaluate(file: UploadFile = File(...)):
    """
    Accepts an Excel file with a 'Prompt' column.
    Generates HTML (via MoEngage Streaming API) and evaluates it.
    Appends results to the ORIGINAL sheet and returns it.
    """
    try:
        # 1. Imports
        import pandas as pd
        import shutil
        import base64
        from io import BytesIO
        import openpyxl
        from openpyxl.drawing.image import Image as ExcelImage
        from openpyxl.utils import get_column_letter
        try:
             from backend.moengage_api import generate_html_from_stream, create_new_session, BYPASS_INSTRUCTION
        except ImportError:
             from moengage_api import generate_html_from_stream, create_new_session, BYPASS_INSTRUCTION

        # 2. Read Valid Excel
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))
        
        if "Prompt" not in df.columns:
             raise HTTPException(status_code=400, detail="Excel must have a 'Prompt' column (Case-sensitive).")
             
        # 3. Define Parallel Worker Function
        async def process_row(index, row):
            prompt = str(row['Prompt'])
            logger.info(f"Starting Row {index+1}: {prompt[:30]}...")
            
            result = {
                "Debug_Raw_Parsing_Prompt": prompt,
                "Debug_Full_API_Prompt": prompt + BYPASS_INSTRUCTION,
                "Debug_API_Response": "",
                "Generated_HTML": "GENERATION_FAILED", # Default
                "Clean_HTML_Output": "GENERATION_FAILED",
                "Score_Fidelity": 0,
                "Score_Visual": 0,
                "Score_Accessibility": 0,
                "Score_Responsiveness": 0,
                "Score_Syntax": 0,
                "Score_Interactive": 0,
                "Verdict": "ERROR",
                "Rationale": "Processing Error",
                "Test_Log": ""
            }
            
            # A. Create Session (Blocking -> Thread)
            try:
                session_id = await run_in_threadpool(create_new_session)
                if not session_id:
                    result["Test_Log"] = "Failed to create session."
                    return result
                
                result["Session_ID"] = session_id
            except Exception as e:
                result["Test_Log"] = f"Session Creation Exception: {e}"
                return result

            # B. Generate HTML (Blocking -> Thread)
            # Pass session_id explicitly
            try:
                html_content, api_log = await run_in_threadpool(generate_html_from_stream, prompt, session_id)
                result["Debug_API_Response"] = api_log
            except Exception as e:
                result["Test_Log"] = f"Generation Exception: {e}"
                return result
                
            if not html_content:
                result["Rationale"] = "Failed to generate HTML from API."
                result["Test_Log"] = f"API Error: {api_log}"
                return result
                
            # Success: Update HTML fields
            result["Generated_HTML"] = html_content
            result["Clean_HTML_Output"] = html_content

            # C. Evaluate (Async)
            try:
                eval_result = await analyze_chat([{"role": "user", "content": html_content}])
                
                result["Score_Fidelity"] = eval_result.score_fidelity
                result["Score_Visual"] = eval_result.score_visual
                result["Score_Accessibility"] = eval_result.score_accessibility
                result["Score_Responsiveness"] = eval_result.score_responsiveness
                result["Score_Syntax"] = getattr(eval_result, 'score_syntax', 0)
                result["Score_Interactive"] = getattr(eval_result, 'score_interactive', 0)
                
                # Pass Base64 strings to result dict (to be handled by Excel Writer)
                result["Screenshot_Portrait"] = getattr(eval_result, 'screenshot_portrait', "")
                result["Screenshot_Landscape"] = getattr(eval_result, 'screenshot_landscape', "")
                
                result["Verdict"] = eval_result.final_judgement
                result["Rationale"] = eval_result.rationale
                result["Test_Log"] = "\n".join(eval_result.execution_trace)
                
            except Exception as e:
                logger.error(f"Row {index+1} Eval Failed: {e}")
                result["Verdict"] = "EVAL_ERROR"
                result["Test_Log"] = f"Evaluation Exception: {e}"
                
            return result

        # 4. Launch Tasks in Parallel (with Semaphore)
        # Limit to 3 concurrent API calls as per user request (Safe Bet)
        sem = asyncio.Semaphore(3)

        async def process_row_with_sem(index, row):
             async with sem:
                 return await process_row(index, row)

        logger.info(f"Processing {len(df)} rows in PARALLEL (Concurrency restricted to 3)...")
        tasks = [process_row_with_sem(i, r) for i, r in df.iterrows()]
        results = await asyncio.gather(*tasks)
        
        # 5. Map Results back to DataFrame Columns
        # results list corresponds to df indices 0..N
        
        # Initialize lists for new columns
        new_cols = {k: [] for k in results[0].keys()}
        
        for res in results:
            for k, v in res.items():
                new_cols[k].append(v)
                
        # Assign to DF
        for k, v in new_cols.items():
            df[k] = v
        
        # 6. Process Images and Return File
        output_stream = BytesIO()
        df.to_excel(output_stream, index=False, engine='openpyxl')
        output_stream.seek(0)
        
        # Load workbook to insert images
        wb = openpyxl.load_workbook(output_stream)
        ws = wb.active
        
        # Find Screenshot Columns
        portrait_col_idx = None
        landscape_col_idx = None
        
        for idx, col in enumerate(df.columns, 1): # 1-based index
            if col == "Screenshot_Portrait":
                portrait_col_idx = idx
            elif col == "Screenshot_Landscape":
                landscape_col_idx = idx
        
        # Insert Images
        if portrait_col_idx or landscape_col_idx:
            # Set generic row height for all data rows
            for row in range(2, len(df) + 2):
                ws.row_dimensions[row].height = 100
                
            if portrait_col_idx:
                 col_letter = get_column_letter(portrait_col_idx)
                 ws.column_dimensions[col_letter].width = 30
                 
            if landscape_col_idx:
                 col_letter = get_column_letter(landscape_col_idx)
                 ws.column_dimensions[col_letter].width = 50

            for row_idx, row_data in df.iterrows():
                excel_row = row_idx + 2 # Header is 1, df is 0-indexed
                
                # Helper to add image
                def add_image_to_cell(b64_str, col_idx):
                    if not b64_str or not isinstance(b64_str, str): return
                    try:
                        img_data = base64.b64decode(b64_str)
                        img = ExcelImage(BytesIO(img_data))
                        # Resize to fit reasonably (e.g., height 120px)
                        # Maintain aspect ratio usually
                        img.height = 120
                        img.width = 120 * (img.width / img.height) 
                        
                        cell = ws.cell(row=excel_row, column=col_idx)
                        ws.add_image(img, cell.coordinate)
                        cell.value = "" # Clear base64 string
                    except Exception as e:
                        print(f"Image Error Row {excel_row}: {e}")

                if portrait_col_idx:
                    add_image_to_cell(row_data.get("Screenshot_Portrait"), portrait_col_idx)
                
                if landscape_col_idx:
                    add_image_to_cell(row_data.get("Screenshot_Landscape"), landscape_col_idx)

        # Save final workbook with images
        final_stream = BytesIO()
        wb.save(final_stream)
        final_stream.seek(0)
        
        from fastapi.responses import StreamingResponse
        headers = {
            'Content-Disposition': 'attachment; filename="evaluated_results_with_screens.xlsx"'
        }
        return StreamingResponse(final_stream, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)
        
    except Exception as e:
        logger.error(f"Batch Processing Error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

from fastapi.staticfiles import StaticFiles

# Create static directory if it doesn't exist (it should)
# Ensure we serve the frontend
# mounted at the end to avoid overriding API routes
# Skip this if running on Vercel, as Vercel handles static files via rewrites
import os
# Static files should be handled by Nginx or Vercel
# app.mount("/", StaticFiles(directory=frontend_path, html=True), name="frontend")

import logging
import sys

# Configure Logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

if __name__ == "__main__":
    import uvicorn
    logger.info("Starting Backend Server...")
    uvicorn.run(app, host="0.0.0.0", port=8000)
